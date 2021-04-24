from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
from functions import *

path_to_file = os.path.join(os.getcwd(), 'docs')
file_to_open = os.path.join(path_to_file, "schet_factura.xlsx")

# стиль текста
title_font = Font(name='Arial', size=12, bold=True, italic=False)
simple_font = Font(name='Arial', size=8, bold=False, italic=False)

# стиль ячеек
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# позиционирование
left_alignment = Alignment(horizontal='left',
                           vertical='center',
                           text_rotation=0,
                           wrap_text=True,
                           shrink_to_fit=False,
                           indent=0)

left_alignment_name = Alignment(horizontal='left',
                                vertical='center',
                                text_rotation=0,
                                wrap_text=False,
                                shrink_to_fit=False,
                                indent=0)

center_alignment = Alignment(horizontal='center',
                             vertical='center',
                             text_rotation=0,
                             wrap_text=False,
                             shrink_to_fit=False,
                             indent=0)


def new_doc(user_data: dict, request_number: int, date_now: str):
    """Создаем новый excel файл с данными"""
    # загрузить таблицу
    wb = load_workbook(file_to_open)
    # получить названия листов в таблице
    sheets_names = wb.sheetnames
    # выбираем первый лист
    sheet = wb[sheets_names[0]]

    sheet['A1'] = f'Счет-фактура № {request_number} от {date_now} г.'
    sheet['A1'].font = title_font

    shop_bin = user_data["shop"].get("Кассовый аппарат", ' ')
    shop_name = user_data["shop"].get("ИП/ТОО", ' ')
    shop_address = user_data["shop"].get("Адрес", ' ')
    sheet['A15'] = f'Грузополучатель: БИН/ИИН {shop_bin}, {shop_name}, г.Уральск, {shop_address}'
    sheet['A18'] = f'БИН/ИИН и адрес местонахождения получателя: БИН/ИИН {shop_bin}, г.Уральск,{shop_address}'

    rows = len(user_data['orders'])

    # переместить строки
    sheet.move_range("A24:K33", rows=rows, cols=0)

    table = sheet["A24": f"K{24 + rows}"]

    table_data = []
    for idd, row in enumerate(user_data['orders']):
        order = [idd + 1, row['Номенклатура'], 'шт.', row['Количество'], row['Цена'], row['Сумма'],
                 None, None, None, None, None]
        table_data.append(order)

    for row_tab, row_data in zip(table, table_data):
        for cell_tab, cell_data in zip(row_tab, row_data):
            sheet[cell_tab.coordinate] = cell_data

    for row in table:
        for cell in row:
            cell.border = thin_border
            cell.font = simple_font

    sheet[f'F{24 + rows}'] = user_data['total_sum']

    sheet[f'H{27 + rows}'] = 'Торговый представитель'
    sheet[f'H{27 + rows}'].alignment = left_alignment_name

    sheet[f'H{30 + rows}'] = user_data['employee']['Сокращенное имя']
    sheet[f'H{30 + rows}'].alignment = left_alignment_name

    for x in range(24, 25 + rows):
        sheet[f'A{x}'].alignment = center_alignment
        sheet[f'B{x}'].alignment = left_alignment
        sheet[f'C{x}'].alignment = center_alignment
        sheet[f'D{x}'].alignment = center_alignment
        sheet[f'E{x}'].alignment = center_alignment
        sheet[f'F{x}'].alignment = center_alignment

    sheet[f'A{24 + rows}'].alignment = left_alignment_name

    save_name = os.path.join(path_to_file, f'Счет-фактура {request_number}.xlsx')
    wb.save(save_name)

# sheet['A27'].border = thin_border

# вставить строку - работает как гавно
# sheet.insert_rows(26)

# ограничить лист - не очень работает
# sheet.ScrollArea = "A1:K37"

# чтение
# print(sheet.max_column)
# print(sheet.max_row)
# print(sheet['A1'].value)
# sheet['A3'] = 1
